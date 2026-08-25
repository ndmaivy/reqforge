from __future__ import annotations

import csv
import io
from datetime import datetime
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.core.exceptions import FeedbackNotFound, ImportFileError, InvalidStateTransition
from app.db.models import Feedback
from app.db.models.enums import FeedbackStatus
from app.modules.feedback.normalizer import normalize_feedback_content
from app.modules.feedback.repository import FeedbackRepository
from app.modules.feedback.schemas import FeedbackCreate, FeedbackUpdate
from app.modules.projects.service import ProjectService


class FeedbackService:
    def __init__(self, session: Session) -> None:
        self.session = session
        self.repository = FeedbackRepository(session)
        self.projects = ProjectService(session)

    def create(
        self, project_id: UUID, payload: FeedbackCreate, owner_id: UUID | None = None
    ) -> Feedback:
        self.projects.get(project_id, owner_id)
        values = payload.model_dump()
        values["content"] = normalize_feedback_content(values["content"])
        feedback = self.repository.create(Feedback(project_id=project_id, **values))
        self.session.commit()
        self.session.refresh(feedback)
        return feedback

    def get(self, feedback_id: UUID, owner_id: UUID | None = None) -> Feedback:
        feedback = self.repository.get(feedback_id)
        if feedback is None:
            raise FeedbackNotFound("Feedback not found")
        if owner_id is not None:
            self.projects.get(feedback.project_id, owner_id)
        return feedback

    def list(
        self,
        project_id: UUID,
        page: int,
        page_size: int,
        status: FeedbackStatus | None = None,
        source: str | None = None,
        category: str | None = None,
        search: str | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        owner_id: UUID | None = None,
    ) -> tuple[list[Feedback], int]:
        self.projects.get(project_id, owner_id)
        return self.repository.list(
            project_id, page, page_size, status, source, category, search, date_from, date_to
        )

    def update(
        self, feedback_id: UUID, payload: FeedbackUpdate, owner_id: UUID | None = None
    ) -> Feedback:
        feedback = self.get(feedback_id, owner_id)
        if feedback.status is FeedbackStatus.ARCHIVED:
            raise InvalidStateTransition("Archived feedback cannot be edited")
        for field, value in payload.model_dump(exclude_unset=True).items():
            if field == "content" and value is not None:
                value = normalize_feedback_content(value)
            setattr(feedback, field, value)
        self.session.commit()
        self.session.refresh(feedback)
        return feedback

    def import_file(
        self,
        project_id: UUID,
        filename: str | None,
        content: bytes,
        owner_id: UUID | None = None,
    ) -> list[Feedback]:
        self.projects.get(project_id, owner_id)
        suffix = Path(filename or "").suffix.lower()
        if suffix == ".csv":
            rows = self._read_csv(content)
        elif suffix == ".xlsx":
            rows = self._read_xlsx(content)
        else:
            raise ImportFileError("Only .csv and .xlsx files are supported")
        if not rows:
            raise ImportFileError("The import file contains no data rows")
        created: list[Feedback] = []
        errors: list[dict[str, object]] = []
        for row_number, row in enumerate(rows, start=2):
            try:
                payload = FeedbackCreate.model_validate(
                    {
                        "content": row.get("content"),
                        "source": row.get("source") or None,
                        "feedback_date": row.get("feedback_date") or None,
                    }
                )
                values = payload.model_dump()
                values["content"] = normalize_feedback_content(values["content"])
                created.append(
                    self.repository.create(Feedback(project_id=project_id, **values))
                )
            except ValidationError as exc:
                errors.append({"row": row_number, "errors": exc.errors()})
        if errors:
            self.session.rollback()
            raise ImportFileError("One or more import rows are invalid", {"rows": errors})
        self.session.commit()
        for item in created:
            self.session.refresh(item)
        return created

    @staticmethod
    def _read_csv(content: bytes) -> list[dict[str, object]]:
        try:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            if reader.fieldnames is None or "content" not in reader.fieldnames:
                raise ImportFileError("Import file must contain a 'content' column")
            return [dict(row) for row in reader]
        except UnicodeDecodeError as exc:
            raise ImportFileError("CSV file must use UTF-8 encoding") from exc

    @staticmethod
    def _read_xlsx(content: bytes) -> list[dict[str, object]]:
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
            if "content" not in headers:
                raise ImportFileError("Import file must contain a 'content' column")
            return [dict(zip(headers, values, strict=False)) for values in iterator]
        except StopIteration:
            return []
        except ImportFileError:
            raise
        except Exception as exc:
            raise ImportFileError("The XLSX file could not be read") from exc

    def archive(self, feedback_id: UUID, owner_id: UUID | None = None) -> Feedback:
        feedback = self.get(feedback_id, owner_id)
        if feedback.status is FeedbackStatus.ARCHIVED:
            raise InvalidStateTransition("Feedback is already archived")
        feedback.status = FeedbackStatus.ARCHIVED
        self.session.commit()
        self.session.refresh(feedback)
        return feedback
